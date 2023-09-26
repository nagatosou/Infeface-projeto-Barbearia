import openpyxl

class backend_menu:
    
    @staticmethod
    def verificar_login(username, password):
        try:
            workbook = openpyxl.load_workbook(r'C:\Users\daniel.reis\Desktop\Meus dados\Machine Learning\usuarios.xlsx')
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if row[0] == username and row[1] == password:
                    return True
        except Exception as e:
            print(f"Erro ao verificar o login: {e}")
        return False

    @staticmethod
    def cadastrar_usuario(username, password):
        try:
            workbook = openpyxl.load_workbook(r'C:\Users\daniel.reis\Desktop\Meus dados\Machine Learning\usuarios.xlsx')
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if row[0] == username:
                    return False

            sheet.append([username, password])
            workbook.save(r'C:\Users\daniel.reis\Desktop\Meus dados\Machine Learning\usuarios.xlsx')
            return True
        except Exception as e:
            print(f"Erro ao cadastrar o usuário: {e}")
        return False

